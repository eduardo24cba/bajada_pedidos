from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

import os

import time

def carga_semana():
    try:
        libro = load_workbook("cargasemana.xlsx")
    except:
        #no esta el archivo
        return False
    
    hoja = libro.active
    datos_carga = {}

    #leemos la hoja excluyendo los titulos
    for campo in list(hoja.values)[1:]:
        datos_carga[campo[2].__str__()]= [campo[3], campo[4], campo[7]]

    return datos_carga

def crear_excel(lista_pedidos):
    lista_columnas = [ chr(i) for i in range(65, 91)]#A, B, C,...
    
    nombre_excel = "pedidos_" + time.strftime("%m%d%Y_%H%M%S") + ".xlsx"

    """
    datos_pedido[0][0] = numero de pedido
    
    datos_solicitante = datos_pedido[0][1][0].split("\n")
    datos_solicitante[1] datos del destinatario
    datos_solicitante[-2] notas que saben incluir a donde va el pedido a veces

    datos_pedido[0][2][0] parte
    datos_pedido[0][2][1] sap
    datos_pedido[0][2][2] descripción
    datos_pedido[0][2][5] cantidad
    """
    
    libro = Workbook()
    hoja_pedido = libro.active
    hoja_pedido.title = 'Pedidos'

    link_volver = nombre_excel + "#Pedidos"

    row_pedidos = 2
    row = 2
    column = 1

    columna_total_pedidos = 3
    row_total_pedidos = 1

    titulo_materiales = ["Código Material", "Código SAP", "Descripción Material", "Es Solución", "Componentes", "Cantidad"]

    hoja_pedido.cell(row=1, column=1).value = "Pedido"

    fill = PatternFill(fill_type="solid", start_color='009999FF', end_color='009999FF')
    fillDatos = PatternFill(fill_type="solid", start_color='009999FF', end_color='00D9F7FA')

    datos_carga = carga_semana()

    for datos_pedido in list(lista_pedidos.values()):

        #creamos la hoja con todos los pedidos
        hoja_pedido.cell(row=row_total_pedidos, column=3).value = datos_pedido[0]

        
        col = hoja_pedido.column_dimensions["C"]
        col.number_format = u'#,##0.00€'
        
        datos_solicitante = datos_pedido[1][0].split("\n")
        
        hoja_pedido.cell(row=row_total_pedidos, column=4).value = datos_solicitante[1]

        #evitamos que el texto sea muy largo
        hoja_pedido.cell(row=row_total_pedidos, column=5).value = "AMBA" if "priorizar este ticket" in datos_solicitante[-2].lower() else datos_solicitante[-2]

        datos_solicitante = datos_pedido[1][1].split("\n")

        #destino
        destino = datos_solicitante[1].lower()
        
        if "cabecera" in destino:
            destino = destino.replace("cabecera", "")
        elif "o&m" in destino:
            destino = destino.replace("o&m", "")
        elif "subcabecera" in destino:
            destino = destino.replace("subcabecera", "")
        elif "retira de almacén" in destino:
            destino = destino.replace("retira de almacén", "")
            
        hoja_pedido.cell(row=row_total_pedidos, column=6).value = destino

        #hoja_pedido["{columna}{fila}".format(columna=lista_columnas[5-1], fila=row_total_pedidos)].alignment =  Alignment(wrap_text=True)

        hoja_pedido["{columna}{fila}".format(columna=lista_columnas[3-1], fila=row_total_pedidos)].fill = fillDatos
        hoja_pedido["{columna}{fila}".format(columna=lista_columnas[4-1], fila=row_total_pedidos)].fill = fillDatos
        hoja_pedido["{columna}{fila}".format(columna=lista_columnas[5-1], fila=row_total_pedidos)].fill = fillDatos
        hoja_pedido["{columna}{fila}".format(columna=lista_columnas[6-1], fila=row_total_pedidos)].fill = fillDatos
        hoja_pedido["{columna}{fila}".format(columna=lista_columnas[7-1], fila=row_total_pedidos)].fill = fillDatos
        hoja_pedido["{columna}{fila}".format(columna=lista_columnas[8-1], fila=row_total_pedidos)].fill = fillDatos
        hoja_pedido["{columna}{fila}".format(columna=lista_columnas[9-1], fila=row_total_pedidos)].fill = fillDatos

        #cantidad
        hoja_pedido.cell(row=row_total_pedidos, column=7).value = "Cantidad"

        key_pedido = None
        
        #buscarmos el numero de pedido en los datos de carga
        for key in datos_carga.keys():
            if datos_pedido[0] in key:
                key_pedido = key
                break

        datos = datos_carga.get(key_pedido)

        #si no existen datos, ese pedido ingreso recientemente o es antiguo.
        if datos:
            #extramos los datos que necesitamos
            reserva, documento, bultos = datos

            #ingresamos esos datos
            hoja_pedido.cell(row=row_total_pedidos, column=8).value = reserva
            hoja_pedido.cell(row=row_total_pedidos, column=9).value = documento
            hoja_pedido.cell(row=row_total_pedidos, column=10).value = bultos
        
        row_total_pedidos += 1

        ###

        #armamos la lista de pedidos que se muestran a la izquierda
        #numero de pedido
        hoja_pedido.cell(row=row_pedidos, column=1).value = datos_pedido[0]

        #link que se crea para acceder a la hoja del pedido generado
        link = nombre_excel + "#" + datos_pedido[0] + "!A1"
        hoja_pedido.cell(row=row_pedidos, column=1).hyperlink = link
        hoja_pedido.cell(row=row_pedidos, column=1).style = "Hyperlink"

        #creamos la hoja con el numero del pedido como nombre
        sheet = libro.create_sheet(title='%s' % datos_pedido[0])
        #numero de pedido
        sheet.cell(row=1, column=1).value = datos_pedido[0]

    
        
        #datos del destino
        for dato in datos_pedido[1]:
            for item in dato.split("\n"):
                
                sheet.cell(row=row, column=column).value = item
                sheet.alignment = Alignment(wrap_text=True)

                if column % 4 == 0:
                    column = 1
                    row += 1
                else:
                    column += 1
            column = 1
            row += 1
            
        row += 1    
        column = 1

        for titulo in titulo_materiales:
            sheet["{columna}{fila}".format(columna=lista_columnas[column-1], fila=row+2)].fill = fill
            sheet.cell(row=row+2, column=column).value = titulo
            sheet.alignment = Alignment(wrap_text=True)#ajustamos el texto
            column += 1

        row += 1
        column = 1

        #materiales que pide
        for materiales in datos_pedido[2]:
            #si el campo esta vacio no lo agregamos
            if materiales:
                hoja_pedido.cell(row=row_total_pedidos, column=columna_total_pedidos).value = materiales
                columna_total_pedidos += 1
            
            sheet.cell(row=row+2, column=column).value = materiales
            column += 1
            
            #creamos una nueva fila y reseteamos la columna
            if column >= 7:
                row +=1
                column = 1
                columna_total_pedidos = 3
                row_total_pedidos += 1

        sheet.cell(row=row+3, column=1, value = "Regresar").hyperlink = link_volver + "!A%d" % row_pedidos
        sheet.cell(row=row+3, column=1).style = "Hyperlink"

        row = 2
        column = 1
        row_pedidos += 1
        columna_total_pedidos = 3
        

    libro.save(filename=nombre_excel)

    return nombre_excel
